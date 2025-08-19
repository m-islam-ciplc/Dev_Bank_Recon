# tally_parser.py

import re
import string
import pandas as pd
from openpyxl import load_workbook
from calendar import month_name


def extract_account_number(metadata):
    acct_pattern = re.compile(r'(\d{8,}|\d{3,}-\d{7,}|\d{4,}-\d{7,})')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        if "unit" in cell.lower():
            continue
        matches = acct_pattern.findall(cell)
        if matches:
            # Normalize: remove dashes
            acct_no = matches[-1].replace("-", "")
            return acct_no, cell, i
    return "", "", None


def extract_bank_code(metadata):
    bank_patterns = {
        "MTB": ["Mutual Trust Bank", "MTB"],
        "MDB": ["Midland Bank", "Midland", "MDB"],
        "PBL": ["Prime Bank", "PBL"],
        "OBL": ["One Bank", "OBL"],
        "EBL": ["Eastern Bank", "EBL"]
    }
    for i, row in metadata.iterrows():
        cell = str(row[0]).lower()
        for code, patterns in bank_patterns.items():
            for pat in patterns:
                if pat.lower() in cell:
                    return code, row[0], i
    return "", "", None


def extract_statement_period(metadata):
    period_pattern = re.compile(
        r'(\d{1,2}-[A-Za-z]{3}-\d{4})\s*to\s*(\d{1,2}-[A-Za-z]{3}-\d{4})')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = period_pattern.search(cell)
        if match:
            return (match.group(1), match.group(2)), cell, i
    return ("", ""), "", None


def extract_unit_name(metadata):
    unit_pattern = re.compile(r'Unit\s*:?[\s)]*([^)]+)')
    for i, row in metadata.iterrows():
        cell = str(row[0])
        match = unit_pattern.search(cell)
        if match:
            return match.group(1).strip(), cell, i
    # Fallback: just use first row as unit name
    return str(metadata.iloc[0, 0]).strip(), metadata.iloc[0, 0], 0


def clean(val):
    return str(val).strip() if val is not None else ""


def deduplicate_row(row, dup_map):
    res = row[:]
    for val, idxs in dup_map.items():
        found = False
        for i in idxs:
            if clean(res[i]) == val:
                if found:
                    res[i] = ""
                else:
                    found = True
    return res


def process_particulars(value):
    if pd.isna(value):
        return ""
    val = str(value).replace('\r\n', '\n').replace('\r', '\n').strip()
    lines = [line.strip() for line in val.split('\n') if line.strip()]
    if not lines:
        return ""
    header = lines[0]
    if len(lines) > 1:
        details = ' '.join(lines[1:]).strip()
        return f"{header}\n{details}"
    match = re.match(r'^([A-Za-z0-9\-/ ]+)[.:,-]\s*(.+)', header)
    if match:
        header_part = match.group(1).strip()
        detail_part = match.group(2).strip()
        return f"{header_part}\n{detail_part}"
    return header


def extract_vendor_updated(particulars):
    if pd.isna(particulars):
        return ""
    val_str = str(particulars).strip()
    lines = [line.strip() for line in val_str.splitlines() if line.strip()]
    if lines and lines[0].lower().replace(" ", "") == "(asperdetails)" and len(lines) > 1:
        val = lines[1]
        match = re.search(r'\b([A-Za-z]+-CE-\d+-\d+-CI)\b', val)
        if match:
            return match.group(1).upper().replace(" ", "")
        match2 = re.search(r'Payable-([^-]+)-ID', val)
        if match2:
            return match2.group(1).strip().upper().replace(" ", "")
        match3 = re.search(
            r'([A-Za-z .&-]+(?:Ltd|Limited))', val, re.IGNORECASE)
        if match3:
            return match3.group(1).strip().upper().replace(" ", "")
        if 'Amount' in val:
            prefix = val.split('Amount')[0]
            chunks = [c.strip() for c in prefix.split('-') if c.strip()]
            if chunks:
                return re.sub(r'[^A-Z0-9]', '', chunks[-1].upper())
        return val.upper().replace(" ", "")
    else:
        val = lines[0] if lines else ""
        val = re.sub(r'^(adv(?:ance)?|ap)[\s\-]*',
                     '', val, flags=re.IGNORECASE)
        val = re.sub(r'^(m[\s\-\/]*s)[\s\-]*', '', val, flags=re.IGNORECASE)
        val = re.split(r'-ID:', val, flags=re.IGNORECASE)[0]
        val = re.sub(r'\band\b', '', val, flags=re.IGNORECASE)
        val = re.sub(f'[{re.escape(string.punctuation)}\s]+', '', val)
        return val.upper()


def parse_tally_file(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # --------- Find header row ---------
    header_keywords = {"Date", "Particulars",
                       "Vch Type", "Vch No.", "Debit", "Credit"}
    header_row_idx = next((i for i, r in enumerate(ws.iter_rows(values_only=True), 1)
                           if header_keywords.issubset({clean(c) for c in r})), None)
    if not header_row_idx:
        wb.close()
        raise ValueError("Header row not found.")

    # --------- Extract metadata for iloc use ---------
    # Build metadata as DataFrame for iloc reference
    metadata_rows = []
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx-1, values_only=True):
        metadata_rows.append([clean(c) for c in row])
    metadata = pd.DataFrame(metadata_rows)

    # --------- Robust metadata extraction ---------
    acct_no, acct_row_val, acct_row = extract_account_number(metadata)
    bank_code, bank_row_val, bank_row = extract_bank_code(metadata)
    (period_start, period_end), period_val, period_row = extract_statement_period(metadata)

    ledger_date = ""
    ledger_year = ""
    if period_start and period_end:
        try:
            first_date = pd.to_datetime(period_start, format="%d-%b-%Y")
            last_date = pd.to_datetime(period_end, format="%d-%b-%Y")
            if first_date.month == last_date.month:
                ledger_date = month_name[first_date.month]
            if first_date.year == last_date.year:
                ledger_year = str(first_date.year)
        except Exception:
            pass

    unit_name, unit_val, unit_row = extract_unit_name(metadata)

    if not bank_code or not acct_no:
        wb.close()
        raise ValueError(
            "Unmapped bank account detected or account cell missing. Update BANK_ACCT_MAP or check metadata."
        )

    # --------- Unmerge all merged cells ---------
    for rng in list(ws.merged_cells.ranges):
        val = ws[rng.coord.split(":")[0]].value
        ws.unmerge_cells(str(rng))
        for row in ws[rng.coord]:
            for cell in row:
                cell.value = val

    headers = [clean(
        c.value) if c.value else f"Unnamed_{i+1}" for i, c in enumerate(ws[header_row_idx])]

    # Rename "Particulars" column to "dr_cr"
    headers = ["dr_cr" if h == "Particulars" and i == headers.index(
        "Particulars") else h for i, h in enumerate(headers)]

    # Rename the next column to 'Particulars'
    # Get the index of the next column after "dr_cr"
    particulars_index = headers.index("dr_cr") + 1
    if particulars_index < len(headers):
        headers[particulars_index] = "Particulars"

    num_cols = len(headers)

    collapsed_rows = []
    current_row = None
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        cleaned = [clean(c) for c in row][:num_cols] + \
            [""] * (num_cols - len(row))
        if (
            (not cleaned[headers.index("Date")] if "Date" in headers else True)
            and (not cleaned[headers.index("dr_cr")] if "dr_cr" in headers else True)
            and (cleaned[headers.index("Particulars")] if "Particulars" in headers else False)
            and current_row is not None
        ):
            idx = headers.index("Particulars")
            current_row[idx] = (current_row[idx] + "\n" + cleaned[idx]).strip()
        else:
            if current_row is not None:
                collapsed_rows.append(current_row)
            current_row = cleaned
    if current_row is not None:
        collapsed_rows.append(current_row)

    wb.close()
    data_rows = collapsed_rows
    dedup_map = {v: idxs for v, idxs in pd.Series(data_rows[0]).groupby(
        lambda x: x).groups.items() if len(idxs) > 1}
    data_rows = [deduplicate_row(row, dedup_map) for row in data_rows]

    if all(clean(v).replace('.', '', 1).replace(',', '', 1).isdigit() or clean(v) == "" for v in data_rows[-1]):
        data_rows.pop(-1)

    df = pd.DataFrame(data_rows, columns=headers).dropna(axis=1, how='all')
    df = df.loc[:, (df != '').any(axis=0)]
    df = df.loc[:, ~df.columns.str.match(r'Unnamed_\d+')]

    # Process "Particulars" column formatting and vendor extraction
    if "Particulars" in df.columns:
        df["Particulars"] = df["Particulars"].apply(process_particulars)
        df["tally_ven"] = df["Particulars"].apply(extract_vendor_updated)
    else:
        df["tally_ven"] = ""

    # Normalize date columns to yyyy-mm-dd format
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(
            df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

    # Remove rows where Particulars == 'Opening Balance' or starts with 'Closing Balance'
    if "Particulars" in df.columns:
        df = df[df["Particulars"].str.strip().str.lower() != "opening balance"]
        df = df[~df["Particulars"].str.strip(
        ).str.lower().str.startswith("closing balance")]

    def to_hex(val):
        try:
            return hex(int(float(val)))[2:]
        except Exception:
            return ""

    uids = []
    rownum = 1
    for i, row in df.iterrows():
        date_val = row.get("Date", "")
        credit_val = row.get("Credit", "")
        debit_val = row.get("Debit", "")
        balance_val = credit_val if (pd.notna(credit_val) and str(
            credit_val).strip() != "") else debit_val
        if pd.notna(date_val) and date_val != "":
            date_str = str(date_val).replace("-", "")
            hexdate = to_hex(date_str)
            try:
                hexbal = to_hex(
                    round(float(str(balance_val).replace(",", ""))))
            except Exception:
                hexbal = ""
            uid = f"T_{bank_code}_{hexdate}_{hexbal}_{rownum:06d}"

            uids.append(uid)
            rownum += 1
        else:
            uids.append("")
    df["tally_uid"] = uids
    cols = ["tally_uid", "bank_code", "acct_no", "unit_name", "statement_month", "statement_year"] + \
        [c for c in df.columns if c not in ["tally_uid", "bank_code",
                                            "acct_no", "unit_name", "statement_month", "statement_year"]]

    df["bank_code"] = bank_code
    # Remove dashes from account number
    df["acct_no"] = acct_no.replace("-", "")
    df["unit_name"] = unit_name
    df["statement_month"] = ledger_date
    df["statement_year"] = ledger_year
    df = df[cols]

    # Optionally: convert empty Debit/Credit to None for DB
    if "Debit" in df.columns:
        df["Debit"] = df["Debit"].apply(
            lambda x: None if str(x).strip() == '' else x)
    if "Credit" in df.columns:
        df["Credit"] = df["Credit"].apply(
            lambda x: None if str(x).strip() == '' else x)

    # Rename columns before writing to df
    new_column_names = {
        "Date": "T_Date",               # Renaming 'Date' to 'T_Date'
        "Particulars": "T_Particulars",  # Newly renamed Particulars to T_Particulars
        "Vch Type": "T_Vch_Type",        # Renaming 'Vch Type' to 'T_Vch_Type'
        "Vch No.": "T_Vch_No",           # Renaming 'Vch No.' to 'T_Vch_No'
        "Debit": "T_Debit",               # Renaming 'Debit' to 'T_Debit'
        "Credit": "T_Credit",             # Renaming 'Credit' to 'T_Credit'
        # Add any other columns you want to rename here
    }

    # Apply the column renaming to df
    df = df.rename(columns=new_column_names)

    # Return the final df with the renamed columns
    return df
